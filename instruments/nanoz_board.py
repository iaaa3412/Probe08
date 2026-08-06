from __future__ import annotations

import csv
import datetime as dt
import json
import os
import re
import shutil
import struct
import threading
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Optional

import serial
from serial.tools import list_ports

BAUD = 921600
READ_TIMEOUT_S = 0.05

VER_RE = re.compile(r"SW:(V[^\s]+).*?S/N:\s*([0-9A-Fa-f]+-[0-9A-Fa-f]+)", re.S)
WHOAMI_RE = re.compile(r"Iam\s+([0-9A-Fa-f]+)", re.I)
ENV_HEADER_RE = re.compile(rb"#env(\d)!\s+(\d+)\s+([0-9A-Fa-f]+)\s+(\d+)\s+(\d+)\s*$", re.I)
SPL_HEADER_RE = re.compile(rb"#spl!\s+(\d+)\s+([0-9A-Fa-f]+)\s+(\d+)\s+(\d+)\s+(\d+)\s*$", re.I)
SEQ_HEADER_RE = re.compile(rb"#seq!\s+(-?\d+)\s+(-?\d+)\s*$", re.I)
EEP_HEADER_RE = re.compile(rb"#eep!\s+(\d+)\s+([0-9A-Fa-f]+)\s+(\d+)\s*$", re.I)


class NanoZError(RuntimeError):
    pass


@dataclass
class PortMeta:
    device: str
    description: str
    hwid: str
    serial_number: str
    vid_pid: str
    location: str


@dataclass
class BoardIdentity:
    port: str
    serial_number: str
    firmware: str
    signature: str
    raw_ver: str
    raw_whoami: str
    usb_id: str
    slot0: Optional[int] = None  # physical probe-head slot (1..N) wired to chip 0
    slot1: Optional[int] = None  # physical probe-head slot (1..N) wired to chip 1
    # Last COM port this board was actually found on, persisted purely as a
    # hint so Connect All can try it directly instead of requiring a full
    # Discover Boards scan every session - NOT identity (see save_known_boards),
    # since Windows can still reassign it; if the board isn't there anymore
    # this hint just fails quietly and the user re-runs Discover Boards.
    last_port: Optional[str] = None

    def chip_slots(self) -> dict:
        return {"0": self.slot0, "1": self.slot1}


def now_stamp() -> str:
    return dt.datetime.now().isoformat(timespec="milliseconds")


def list_serial_ports() -> list[PortMeta]:
    out: list[PortMeta] = []
    for p in sorted(list_ports.comports(), key=lambda x: x.device):
        vid_pid = ""
        if p.vid is not None and p.pid is not None:
            vid_pid = f"{p.vid:04X}:{p.pid:04X}"
        out.append(
            PortMeta(
                device=p.device,
                description=p.description or "",
                hwid=p.hwid or "",
                serial_number=p.serial_number or "",
                vid_pid=vid_pid,
                location=p.location or "",
            )
        )
    return out


def open_serial(port: str) -> serial.Serial:
    ser = serial.Serial(
        port=port,
        baudrate=BAUD,
        bytesize=serial.EIGHTBITS,
        parity=serial.PARITY_NONE,
        stopbits=serial.STOPBITS_ONE,
        timeout=READ_TIMEOUT_S,
        write_timeout=1.0,
        xonxoff=False,
        rtscts=False,
        dsrdtr=False,
    )
    time.sleep(0.2)
    ser.reset_input_buffer()
    ser.reset_output_buffer()
    return ser


def send_ascii(ser: serial.Serial, cmd: str) -> None:
    if not cmd.endswith("\r"):
        cmd += "\r"
    ser.write(cmd.encode("ascii"))
    ser.flush()


def read_text_for(ser: serial.Serial, seconds: float) -> str:
    deadline = time.time() + seconds
    data = bytearray()
    while time.time() < deadline:
        chunk = ser.read(4096)
        if chunk:
            data.extend(chunk)
        else:
            time.sleep(0.01)
    return data.decode(errors="replace").strip()


def identify_on_port(port: str) -> Optional[BoardIdentity]:
    ser = open_serial(port)
    try:
        send_ascii(ser, "ver")
        raw_ver = read_text_for(ser, 0.75)
        send_ascii(ser, "whoami")
        raw_whoami = read_text_for(ser, 0.50)
    except Exception:
        return None
    finally:
        ser.close()

    m = VER_RE.search(raw_ver)
    if not m:
        return None

    firmware, sn = m.group(1), m.group(2).upper()
    wm = WHOAMI_RE.search(raw_whoami)
    signature = wm.group(1) if wm else ""

    usb_id = ""
    for p in list_serial_ports():
        if p.device.upper() == port.upper():
            usb_id = p.serial_number or f"VIDPID={p.vid_pid};LOC={p.location};PORT={p.device}"
            break

    return BoardIdentity(
        port=port,
        serial_number=sn,
        firmware=firmware,
        signature=signature,
        raw_ver=raw_ver,
        raw_whoami=raw_whoami,
        usb_id=usb_id,
    )


def discover_boards(ports: Optional[list[str]] = None,
                    log: Optional[Callable[[str], None]] = None) -> list[BoardIdentity]:
    candidates = ports if ports is not None else [p.device for p in list_serial_ports()]
    found: list[BoardIdentity] = []
    for port in candidates:
        if log:
            log(f"Probing {port}...")
        try:
            ident = identify_on_port(port)
        except Exception as e:
            if log:
                log(f"  -> could not open ({e}) - in use by another program?")
            continue
        if ident:
            found.append(ident)
            if log:
                log(f"  -> NanoZ board found: S/N {ident.serial_number}  FW {ident.firmware}")
        elif log:
            log(f"  -> no response (not a NanoZ board, or powered off)")
    return found


def read_line_bytes(ser: serial.Serial, buffer: bytearray, timeout_s: float = 2.0) -> Optional[bytes]:
    deadline = time.time() + timeout_s
    while time.time() < deadline:
        idx = buffer.find(b"\n")
        if idx >= 0:
            line = bytes(buffer[:idx]).strip(b"\r\n ")
            del buffer[: idx + 1]
            if line:
                return line
            continue
        chunk = ser.read(4096)
        if chunk:
            buffer.extend(chunk)
        else:
            time.sleep(0.005)
    return None


def read_exact_from_buffer(ser: serial.Serial, buffer: bytearray, n: int, timeout_s: float = 2.0) -> bytes:
    deadline = time.time() + timeout_s
    while len(buffer) < n and time.time() < deadline:
        chunk = ser.read(n - len(buffer))
        if chunk:
            buffer.extend(chunk)
        else:
            time.sleep(0.005)
    if len(buffer) < n:
        raise NanoZError(f"Timed out waiting for binary block: needed {n}, got {len(buffer)}")
    data = bytes(buffer[:n])
    del buffer[:n]
    return data


def parse_spl_data(data: bytes) -> dict:
    if len(data) < 48:
        raise NanoZError(f"SPL data block too short: {len(data)} bytes")
    vals = struct.unpack_from("<IBBH4h4f4f", data, 0)
    ppms, chip_id, sensor_mask, reserved = vals[:4]
    dac = vals[4:8]
    adc = vals[8:12]
    heaters = vals[12:16]
    return {
        "ppms": ppms,
        "chip_id": chip_id,
        "sensor_mask": sensor_mask,
        "reserved": reserved,
        "dac_mv_s1": dac[0],
        "dac_mv_s2": dac[1],
        "dac_mv_s3": dac[2],
        "dac_mv_s4": dac[3],
        "adc_current_ma_s1": adc[0],
        "adc_current_ma_s2": adc[1],
        "adc_current_ma_s3": adc[2],
        "adc_current_ma_s4": adc[3],
        "heater1_voltage_mv": heaters[0],
        "heater1_current_ma": heaters[1],
        "heater2_voltage_mv": heaters[2],
        "heater2_current_ma": heaters[3],
    }


def parse_env_data(data: bytes) -> dict:
    if len(data) < 132:
        raise NanoZError(f"ENV data block too short: {len(data)} bytes")
    off = 0
    pps, samples_nb, adc_mask = struct.unpack_from("<IHH", data, off)
    off += 8
    adc_samples = struct.unpack_from("<8H", data, off)
    off += 16
    adc_voltage = struct.unpack_from("<8f", data, off)
    off += 32
    adc_current = struct.unpack_from("<8f", data, off)
    off += 32
    htr_voltage = struct.unpack_from("<4f", data, off)
    off += 16
    adc_mid, mcu_temp = struct.unpack_from("<ff", data, off)
    off += 8
    humidity_x100, tempH_x100, pressure_x10, tempP_x100, pending, align = struct.unpack_from("<6h", data, off)
    off += 12
    age = struct.unpack_from("<2I", data, off)

    return {
        "pps": pps,
        "adc_samples_nb": samples_nb,
        "adc_mask": adc_mask,
        "adc_mid_value": adc_mid,
        "mcu_temperature_c": mcu_temp,
        "humidity_percent": humidity_x100 / 100.0,
        "temp_h_c": tempH_x100 / 100.0,
        "pressure_hpa_minus_1013": pressure_x10 / 10.0,
        "temp_p_c": tempP_x100 / 100.0,
        "pending": pending,
        "align": align,
        "age_chip1_s": age[0],
        "age_chip2_s": age[1],
        "adc_samples_4x2": ";".join(str(x) for x in adc_samples),
        "adc_voltage_4x2": ";".join(f"{x:.6g}" for x in adc_voltage),
        "adc_current_4x2": ";".join(f"{x:.6g}" for x in adc_current),
        "htr_voltage_2x2": ";".join(f"{x:.6g}" for x in htr_voltage),
    }


# EEPROM "Configuration" layout — reverse-engineered live against a real
# EK-IV board and Nanoz_EK.exe (2026-08-04), NOT from any vendor
# documentation (neither reference PDF documents this). Nanoz_EK.exe embeds
# named constants (EEPROM_PARAMS_ADDR, EEPROM_CYCLES_PAGE, etc.) via its own
# Free Pascal RTTI/debug info, but those ADDR constants turned out to be
# red herrings — the real byte address of a section is PAGE * PAGE_SIZE, not
# the ADDR constant itself. Confirmed by diffing a real Sequence Duration
# edit (5s -> 2s) against a live rdeep and finding it land exactly at the
# predicted offset. Cycle-record layout confirmed against 2 real cycles
# (32 bytes apart, matching CYCLE_RECORD_SIZE). Sequence-record layout is
# only confirmed for Duration (offset +2) and Chip (offset +54) - only ONE
# real sequence has ever been observed, so the exact stride between
# consecutive sequence records, and the sub-order of the heater ramp
# fields, are NOT yet confirmed. Read-only so far - no wreep support, since
# writing to an unconfirmed offset could corrupt the board's stored config
# with no way to detect or undo it (see gui/nanoz_panel.py's NanoZ_EK tab).
EEPROM_PAGE_SIZE = 32
EEPROM_PARAMS_ADDR = 0
EEPROM_CYCLES_PAGE = 16
EEPROM_CYCLES_ADDR = EEPROM_CYCLES_PAGE * EEPROM_PAGE_SIZE       # 512
EEPROM_SEQUENCES_PAGE = 64
EEPROM_SEQUENCES_ADDR = EEPROM_SEQUENCES_PAGE * EEPROM_PAGE_SIZE  # 2048
EEPROM_CYCLE_RECORD_SIZE = 32
MAX_CYCLES_NB = 48
MAX_SEQUENCE_NB = 96


def parse_params_block(data: bytes) -> dict:
    """Decode the EEPROM_PARAMS_ADDR (0) region: device signature, cycle
    count, periodicity, the board's own CAL-1/CAL-2 calibration offsets
    (same values `calib ?` reports), and the two installed chips'
    identity/age records. Confirmed field-for-field against Nanoz_EK.exe's
    own display (Signature, CAL-1/CAL-2, and the "ID:"/"Age:" fields, whose
    "D{W}L{X}-{Y}-{Z}" format matches this decode's w/x/y/z exactly)."""
    if len(data) < 152:
        raise NanoZError(f"PARAMS block too short: {len(data)} bytes (need >= 152)")
    signature = struct.unpack_from("<H", data, 0)[0]
    cycles_configured = struct.unpack_from("<H", data, 4)[0]
    periodicity_ms = data[14]
    cal1, cal2 = struct.unpack_from("<ff", data, 20)

    def chip_record(off):
        w, x, y, z, age_s = struct.unpack_from("<5I", data, off)
        return {"w": w, "x": x, "y": y, "z": z, "age_s": age_s,
               "id": f"D{w}L{x}-{y}-{z}"}

    return {
        "signature": f"0x{signature:04X}",
        "cycles_configured": cycles_configured,
        "periodicity_ms": periodicity_ms,
        "cal1": cal1,
        "cal2": cal2,
        "chip1": chip_record(112),
        "chip2": chip_record(132),
    }


def parse_cycle_record(data: bytes) -> "dict | None":
    """Decode one EEPROM_CYCLE_RECORD_SIZE-byte (32) cycle record. Returns
    None if it's erased/unused (all 0xFF). wire_index is 0-based - Nanoz_EK's
    UI "Cycle 1"/"Cycle 2" are wire index 0/1 (same off-by-one as `run <nn>`,
    confirmed against the manual's error text for an out-of-range index)."""
    if len(data) < EEPROM_CYCLE_RECORD_SIZE:
        raise NanoZError(f"Cycle record too short: {len(data)} bytes")
    if all(b == 0xFF for b in data[:EEPROM_CYCLE_RECORD_SIZE]):
        return None
    wire_index, num_sequences = struct.unpack_from("<HH", data, 0)
    seq_refs = []
    off = 4
    for _ in range(min(num_sequences, (EEPROM_CYCLE_RECORD_SIZE - 4) // 4)):
        seq_refs.append(struct.unpack_from("<I", data, off)[0])
        off += 4
    return {"wire_index": wire_index, "num_sequences": num_sequences,
           "sequence_refs": seq_refs}


# Named-field byte offsets within a sequence record, mapped against the
# real Nanoz_EK.exe field names from references/250723_User manual EK IV.pdf
# section IV.D (Sequence: D.a "Sequence settings", D.b "Heater settings" /
# Table 2). Values matched byte-for-byte against the UI for one real,
# fully-populated sequence (2026-08-05) - see parse_sequence_records'
# docstring for the full derivation and confidence notes per field. This
# dict exists so a future `wreep`-based write can target the same offsets
# used here for reading, without re-deriving them.
SEQ_FIELD_OFFSETS = {
    "duration_s": 2,        # CONFIRMED (live diff, 5s -> 2s)
    "delay_s": 4,            # matches UI "Delay", not independently diffed
    "sensor_mv": 6,           # manual: "one voltage applied for all sensors" -
                              # UI has a single Sensors-NZG2 field; the board
                              # stores 4 (offsets 6/10/14/18, all equal here)
    "ramp_up_ms": 22,         # Table 2 row 3
    "high_duration_ms": 26,   # Table 2 row 4
    "ramp_down_ms": 30,       # Table 2 row 5
    "low_duration_ms": 34,    # Table 2 row 6 - best-effort pairing, see docstring
    "phase_shift_ms": 38,     # Table 2 row 7 - best-effort pairing, see docstring
    "heater1_low_mv": 42,     # Table 2 row 1 (low state)
    "heater2_low_mv": 46,     # Table 2 row 2 (low state)
    "heater1_high_mv": 50,    # Table 2 row 1 (high state)
    "heater2_high_mv": 54,    # Table 2 row 2 (high state)
    "chip": 58,                # ambiguous vs offset 60, see docstring
    "resolution_ms": 62,       # Table 2 row 8 - ambiguous vs offset 64
}


def parse_sequence_records(data: bytes) -> list:
    """Scan the EEPROM_SEQUENCES_ADDR region for sequence records, each
    terminated by a 0xFFFF marker. Layout confirmed 2026-08-04/05 via a
    live diff against a real, fully-populated sequence (Sensors 1-4=900mV,
    Heater times=100/200/300, Heater extra=400/500, Heater voltages=
    1600/1700/1800/1900, Resolution=0 - every value below was cross-checked
    byte-for-byte against Nanoz_EK.exe's own Sequence/Heater dialog for
    this exact sequence). Field NAMES below (as opposed to raw offsets)
    come from references/250723_User manual EK IV.pdf section IV.D, Table 2
    "Heater control parameters":

    offset 0            u16  wire_index (0-based)
    offset 2   duration_s     i16   CONFIRMED (earlier diff, 5s -> 2s)
    offset 4   delay_s        i16   matches UI "Delay: 0", not independently diffed
    offset 6,10,14,18   i16 x4  sensor_mv/sensors_mv (Sensor 1-4) CONFIRMED
                       value-match, though the manual says the real UI only
                       exposes ONE "Sensors-NZG2" voltage applied to all
                       sensors - offset 6 is treated as that canonical
                       field. Each real value is immediately followed by a
                       constant-800 int16 at +8/+12/+16/+20 whose meaning
                       is still unknown (returned as sensors_pad_raw).
    offset 22 ramp_up_ms, 26 high_duration_ms, 30 ramp_down_ms   i16 x3
                       (the UI's 3 "Times in ms" boxes) CONFIRMED value
                       match against Table 2 rows 3/4/5 by process of
                       elimination (3 values, 3 remaining un-matched Table-2
                       time rows before Low state duration/Phase shift) -
                       each followed by a constant-0 int16 (unused/reserved)
    offset 34 low_duration_ms, 38 phase_shift_ms   i16 x2 (the UI's 2
                       bottom boxes, mislabeled "Voltages in mV" in the
                       dialog but drawn as horizontal ms-style double-
                       arrows) - values match Table 2 rows 6/7 by
                       elimination, this specific pairing/order is a
                       best-effort guess, not yet isolated by its own diff.
                       followed by constant-0 padding.
    offset 42 heater1_low_mv, 46 heater2_low_mv, 50 heater1_high_mv,
    54 heater2_high_mv   i16 x4 (order found on the wire: 1600,1800,1700,
                       1900 - i.e. NOT left-to-right as drawn; grouped as
                       [H1_low, H2_low, H1_high, H2_high] since 1600<1700
                       and 1800<1900, matching Table 2 rows 1/2's "low &
                       high states") CONFIRMED value match - each followed
                       by a constant-2000 int16, meaning unknown (returned
                       as heater_v_pad_raw)
    offset 58, 60      i16, i16  both =1 in this sample - one of these is
                       almost certainly Chip (a different dialog for this
                       same sequence showed "Chip: 1"), but with both equal
                       to 1 there's no way yet to tell which is Chip vs an
                       unrelated flag. NOT the old +54 guess - that offset
                       is actually heater1_high_mv now that a real,
                       non-erased sequence has been observed; +54 only
                       looked like Chip=1 before because the old baseline
                       record's later bytes hadn't been written yet.
    offset 62, 64      i16, i16  both =0 - resolution_ms (Table 2 row 8,
                       "Time resolution", matches UI's "Resolution: Time: 0
                       ms") is presumably one of these, unconfirmed which.
    offset 66          0xFFFF terminator

    Everything above except duration_s is unconfirmed-by-elimination only
    (matched by value equality/count against one real sample and the
    manual's Table 2, not yet isolated by changing that one field alone
    and re-diffing) - treat sensor/heater-times/heater-voltages as high-
    confidence but chip/resolution candidates as genuinely ambiguous, and
    the low_duration/phase_shift and heater-low/high pairings as
    best-effort, until a future diff isolates them."""
    records = []
    start = 0
    n = len(data)
    while start < n:
        if data[start] == 0xFF and (start + 1 >= n or data[start + 1] == 0xFF):
            break  # ran into erased/unused space - no more records
        term = data.find(b"\xff\xff", start)
        if term == -1:
            end = n
        else:
            end = term + 2
        record = data[start:end]
        if len(record) >= 4:
            wire_index = struct.unpack_from("<H", record, 0)[0]
            duration_s = struct.unpack_from("<h", record, 2)[0]
            delay_s = struct.unpack_from("<h", record, 4)[0] if len(record) >= 6 else None
            sensors_mv = sensors_pad_raw = None
            if len(record) >= 22:
                sensors_mv = [struct.unpack_from("<h", record, o)[0] for o in (6, 10, 14, 18)]
                sensors_pad_raw = [struct.unpack_from("<h", record, o)[0] for o in (8, 12, 16, 20)]
            heater_times_ms = None
            if len(record) >= 32:
                heater_times_ms = [struct.unpack_from("<h", record, o)[0] for o in (22, 26, 30)]
            heater_extra = None
            if len(record) >= 40:
                heater_extra = [struct.unpack_from("<h", record, o)[0] for o in (34, 38)]
            heater_voltages_mv = heater_v_pad_raw = None
            if len(record) >= 58:
                heater_voltages_mv = [struct.unpack_from("<h", record, o)[0] for o in (42, 46, 50, 54)]
                heater_v_pad_raw = [struct.unpack_from("<h", record, o)[0] for o in (44, 48, 52, 56)]
            chip_candidates = None
            if len(record) >= 62:
                chip_candidates = [struct.unpack_from("<h", record, o)[0] for o in (58, 60)]
            resolution_candidates = None
            if len(record) >= 66:
                resolution_candidates = [struct.unpack_from("<h", record, o)[0] for o in (62, 64)]
            ht = heater_times_ms or [None, None, None]
            he = heater_extra or [None, None]
            hv = heater_voltages_mv or [None, None, None, None]
            records.append({
                "blob_offset": start,
                "record_len": len(record),
                "wire_index": wire_index,
                "duration_s": duration_s,
                "delay_s": delay_s,
                "sensor_mv": sensors_mv[0] if sensors_mv else None,
                "sensors_mv": sensors_mv,
                "sensors_pad_raw": sensors_pad_raw,
                "ramp_up_ms": ht[0],
                "high_duration_ms": ht[1],
                "ramp_down_ms": ht[2],
                "low_duration_ms": he[0],
                "phase_shift_ms": he[1],
                "heater1_low_mv": hv[0],
                "heater2_low_mv": hv[1],
                "heater1_high_mv": hv[2],
                "heater2_high_mv": hv[3],
                "heater_times_ms": heater_times_ms,
                "heater_extra": heater_extra,
                "heater_voltages_mv": heater_voltages_mv,
                "heater_v_pad_raw": heater_v_pad_raw,
                "chip_candidates": chip_candidates,
                "resolution_candidates": resolution_candidates,
                "chip": chip_candidates[0] if chip_candidates else None,
                "resolution_ms": resolution_candidates[0] if resolution_candidates else None,
                "raw_hex": record.hex(),
            })
        start = end
    return records


def encode_sequence_patch(original_record: bytes, fields: dict) -> bytearray:
    """Patch a real sequence record (as returned by rdeep/parse_sequence_records)
    with new values for the D-portion (Sequence settings + Heater settings)
    fields ONLY - every other byte (wire_index, the sensor/heater padding
    int16s, the ambiguous Chip/Resolution offsets, and the 0xFFFF
    terminator) is left exactly as read. Never construct a record from
    scratch - always patch a real one, since large parts of the layout
    (padding meaning, whether stride varies) are not understood well
    enough to safely regenerate from nothing.

    `fields` keys (all optional - omitted keys keep their original byte
    value): duration_s, delay_s, sensor_mv (single value, written to all 4
    real sensor slots per the manual's "one voltage for all sensors"),
    ramp_up_ms, high_duration_ms, ramp_down_ms, low_duration_ms,
    phase_shift_ms, heater1_low_mv, heater2_low_mv, heater1_high_mv,
    heater2_high_mv. chip/resolution_ms are deliberately NOT accepted here
    - their byte offset is still ambiguous (see parse_sequence_records'
    docstring) and a wrong guess could silently overwrite an unrelated,
    still-unknown field with no checksum-level way to detect it.

    Duration/Delay/heater timing/heater voltage fields are packed as
    UNSIGNED 16-bit (manual's documented ranges go up to 60000, which
    overflows a signed int16) - sensor_mv is packed SIGNED since the
    manual documents a negative sensor bias range (NZGS2: +/-0.8V)."""
    buf = bytearray(original_record)

    def put_u(offset, value):
        struct.pack_into("<H", buf, offset, int(value))

    def put_s(offset, value):
        struct.pack_into("<h", buf, offset, int(value))

    if "duration_s" in fields:
        put_u(SEQ_FIELD_OFFSETS["duration_s"], fields["duration_s"])
    if "delay_s" in fields:
        put_u(SEQ_FIELD_OFFSETS["delay_s"], fields["delay_s"])
    if "sensor_mv" in fields:
        for o in (6, 10, 14, 18):
            put_s(o, fields["sensor_mv"])
    for key in ("ramp_up_ms", "high_duration_ms", "ramp_down_ms", "low_duration_ms",
               "phase_shift_ms", "heater1_low_mv", "heater2_low_mv",
               "heater1_high_mv", "heater2_high_mv"):
        if key in fields:
            put_u(SEQ_FIELD_OFFSETS[key], fields[key])
    return buf


def append_csv_row(path, row: dict) -> None:
    path = Path(path)
    exists = path.exists() and path.stat().st_size > 0
    with path.open("a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(row.keys()))
        if not exists:
            writer.writeheader()
        writer.writerow(row)


BOARDS_MEMORY_FILENAME = "ata_nanoz_boards.json"


def save_known_boards(folder, identities: list) -> None:
    """Persist known boards keyed by serial number, NOT COM port - a board's
    port is assigned by Windows on connect and can (and does) change across
    replugs/reboots, so it's not a stable identity and isn't saved. The
    current port IS saved separately as "last_port" though - purely a hint
    (see BoardIdentity.last_port), not identity. Dedupes by serial_number
    defensively (last one wins) so a transient in-memory duplicate never
    gets written twice."""
    by_sn: dict[str, dict] = {}
    for i in identities:
        by_sn[i.serial_number or f"(no S/N) {i.port}"] = {
            "serial_number": i.serial_number, "firmware": i.firmware,
            "signature": i.signature, "usb_id": i.usb_id, "slot0": i.slot0, "slot1": i.slot1,
            "last_port": i.port or i.last_port or None,
        }
    path = Path(folder) / BOARDS_MEMORY_FILENAME
    path.write_text(json.dumps(list(by_sn.values()), indent=2), encoding="utf-8")


def load_known_boards(folder) -> list[BoardIdentity]:
    """Returns each known board with port="" - its real, LIVE port (if any)
    is only confirmed by actually finding it this session (Discover Boards,
    or Connect All trying last_port directly); see save_known_boards for why
    port itself isn't persisted, only last_port as a hint."""
    path = Path(folder) / BOARDS_MEMORY_FILENAME
    if not path.is_file():
        return []
    try:
        rows = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return []
    return [
        BoardIdentity(
            port="", serial_number=row.get("serial_number", ""),
            firmware=row.get("firmware", ""), signature=row.get("signature", ""),
            raw_ver="", raw_whoami="", usb_id=row.get("usb_id", ""),
            # Legacy single-slot files (pre-two-chip-per-board) had "slot" -> migrate to slot0.
            slot0=row.get("slot0", row.get("slot")),
            slot1=row.get("slot1"),
            last_port=row.get("last_port"),
        )
        for row in rows
    ]


WAFER_PLAN_XLSX_FILENAME = "ata_nanoz_wafer_plan.xlsx"


def wafer_plan_path_in_folder(folder) -> str:
    """The fixed path a wafer plan .xlsx lives at once imported into this ATA
    folder - see import_wafer_plan_into_folder. Whatever the user originally
    picked from disk (e.g. references/nautilusprobeplan.xlsx, which is only
    an example/template) is copied here so the folder is self-contained and
    doesn't depend on that source file still existing at its original path."""
    return str(Path(folder) / WAFER_PLAN_XLSX_FILENAME)


def import_wafer_plan_into_folder(folder, source_path: str) -> str:
    """Copies source_path into the ATA folder at its fixed name (overwriting
    any previous import) and returns that new path. Does not parse it -
    caller should load_wafer_plan the returned path to validate/use it."""
    dest = wafer_plan_path_in_folder(folder)
    if os.path.abspath(source_path) != os.path.abspath(dest):
        shutil.copyfile(source_path, dest)
    return dest


LEGACY_RECIPE_FILENAME = "ata_nanoz_recipe.json"
RECIPES_FILENAME = "ata_nanoz_recipes.json"

_RECIPE_SHOT_META_KEYS = ("die_column", "td_start_row", "td_end_row", "board_reasons",
                          "chip_reasons")


def _shots_to_rows(shots: list) -> list:
    rows = []
    for s in shots:
        row = {"label": s.get("label", ""), "excluded_boards": sorted(s.get("excluded_boards", ()))}
        for k in _RECIPE_SHOT_META_KEYS:
            if k in s:
                row[k] = s[k]
        rows.append(row)
    return rows


def _rows_to_shots(rows: list) -> list[dict]:
    shots = []
    for s in rows:
        if not isinstance(s, dict):
            continue
        shot = {"label": s.get("label", ""), "excluded_boards": set(s.get("excluded_boards", ()))}
        for k in _RECIPE_SHOT_META_KEYS:
            if k in s:
                shot[k] = s[k]
        shots.append(shot)
    return shots


def _load_recipes_file(folder) -> dict:
    path = Path(folder) / RECIPES_FILENAME
    if path.is_file():
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(data.get("recipes"), dict):
                data.setdefault("active", None)
                return data
        except (OSError, ValueError):
            pass
    return {"active": None, "recipes": {}}


def _write_recipes_file(folder, data: dict) -> None:
    path = Path(folder) / RECIPES_FILENAME
    path.write_text(json.dumps(data, indent=2), encoding="utf-8")


def list_recipe_names(folder) -> list[str]:
    return sorted(_load_recipes_file(folder)["recipes"].keys())


def get_active_recipe_name(folder):
    return _load_recipes_file(folder).get("active")


def save_named_recipe(folder, name: str, shots: list, wafer_plan_path: str | None = None) -> None:
    data = _load_recipes_file(folder)
    data["recipes"][name] = _shots_to_rows(shots)
    data["active"] = name
    if wafer_plan_path:
        data.setdefault("wafer_plan_paths", {})[name] = wafer_plan_path
    _write_recipes_file(folder, data)


def load_named_recipe(folder, name: str) -> list[dict]:
    rows = _load_recipes_file(folder)["recipes"].get(name)
    return _rows_to_shots(rows) if rows is not None else []


def get_recipe_wafer_plan_path(folder, name: str) -> str | None:
    """The source .xlsx path a named recipe's shots were generated from, if any
    was recorded — lets the GUI auto-reload the wafer map alongside the recipe."""
    return _load_recipes_file(folder).get("wafer_plan_paths", {}).get(name)


def set_active_recipe(folder, name: str) -> None:
    data = _load_recipes_file(folder)
    if name in data["recipes"]:
        data["active"] = name
        _write_recipes_file(folder, data)


def delete_named_recipe(folder, name: str) -> None:
    data = _load_recipes_file(folder)
    data["recipes"].pop(name, None)
    data.get("wafer_plan_paths", {}).pop(name, None)
    if data.get("active") == name:
        data["active"] = None
    _write_recipes_file(folder, data)


def load_active_recipe(folder):
    """Returns (name, shots, wafer_plan_path) for the last-saved/loaded recipe,
    or (None, [], None)."""
    data = _load_recipes_file(folder)
    name = data.get("active")
    if not name or name not in data["recipes"]:
        return None, [], None
    wafer_plan_path = data.get("wafer_plan_paths", {}).get(name)
    return name, _rows_to_shots(data["recipes"][name]), wafer_plan_path


def migrate_legacy_recipe(folder):
    """One-time migration of the old single-recipe file (pre-naming) into the
    named scheme, under the name 'Imported'. Returns the name if it migrated
    something, else None. No-op if any named recipe already exists."""
    legacy_path = Path(folder) / LEGACY_RECIPE_FILENAME
    if not legacy_path.is_file() or _load_recipes_file(folder)["recipes"]:
        return None
    try:
        legacy = json.loads(legacy_path.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return None
    shots = _rows_to_shots(legacy.get("shots", []))
    if not shots:
        return None
    save_named_recipe(folder, "Imported", shots)
    return "Imported"


# ── Wafer-plan (.xlsx) import ───────────────────────────────────────────────
# Parses a Nautilus-style wafer-plan workbook - "Die Map" (row/col grid of
# die serials; fill color marks product vs reference/skip-test) and
# "Touchdown List" (flat list of Die IDs - the top die of each touchdown, in
# order) sheets - into the geometry needed to auto-generate the NanoZ recipe:
# which of the probe head's slots (1..probe_height, top to bottom) land on a
# real product die, a reference/monitor die, or off the wafer entirely, for
# every touchdown. A third "Probe Overlay" sheet exists in the workbook too,
# but it's a human-readable visual (BOLD marks touchdown starts) generated
# by the same macro that produces Touchdown List - since Touchdown List is
# already that macro's computed result, Probe Overlay isn't parsed here.

try:
    import openpyxl
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

# Die Map fill color for a reference/monitor (skip-test) die - everything
# else with a die serial in it is a normal product die.
_REFERENCE_FILL_RGBS = frozenset({"FFC00000"})

# The probe head's physical slot count (1-20, top to bottom) - a hardware
# constant of the probe card, not something the wafer-plan workbook carries.
DEFAULT_PROBE_HEIGHT = 20


@dataclass
class WaferPlan:
    dies: dict          # (row, col) -> {"serial": str, "status": "product"|"reference"}
    serial_to_rc: dict  # serial.upper() -> (row, col)
    touchdowns: list     # [(row, col), ...] top die of each touchdown, sheet order
    probe_height: int = DEFAULT_PROBE_HEIGHT


def load_wafer_plan(path) -> WaferPlan:
    if not _OPENPYXL_AVAILABLE:
        raise NanoZError("openpyxl is required to import a wafer plan .xlsx (pip install openpyxl)")

    wb = openpyxl.load_workbook(path, data_only=True)
    for name in ("Die Map", "Touchdown List"):
        if name not in wb.sheetnames:
            raise NanoZError(f"'{name}' sheet not found — not a recognized wafer-plan workbook.")

    die_ws = wb["Die Map"]
    dies: dict[tuple[int, int], dict] = {}
    serial_to_rc: dict[str, tuple[int, int]] = {}
    # Row 1 is a title, row 2 is the "row\col" header, data starts row 3;
    # column A holds the row-number label, die data starts column B.
    for row in die_ws.iter_rows(min_row=3, min_col=2):
        for cell in row:
            if not cell.value:
                continue
            serial = str(cell.value).strip()
            r, c = cell.row - 2, cell.column - 1
            fill = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
            status = "reference" if fill in _REFERENCE_FILL_RGBS else "product"
            dies[(r, c)] = {"serial": serial, "status": status}
            serial_to_rc[serial.upper()] = (r, c)
    if not dies:
        raise NanoZError("Die Map: no dies found.")

    td_ws = wb["Touchdown List"]
    touchdowns = []
    missing = []
    # Row 1 is a title, row 2 is the "Die ID" header, data starts row 3.
    for row in td_ws.iter_rows(min_row=3, max_col=1, values_only=True):
        serial = row[0] if row else None
        if not serial:
            continue
        serial = str(serial).strip()
        rc = serial_to_rc.get(serial.upper())
        if rc is None:
            missing.append(serial)
            continue
        touchdowns.append(rc)
    if not touchdowns:
        raise NanoZError("Touchdown List: no touchdown dies found.")
    if missing:
        raise NanoZError(
            f"Touchdown List references {len(missing)} die ID(s) not found on Die Map: "
            + ", ".join(missing[:5]) + (", ..." if len(missing) > 5 else ""))

    return WaferPlan(dies=dies, serial_to_rc=serial_to_rc, touchdowns=touchdowns)


def classify_die(plan: "WaferPlan", row: int, col: int,
                 row_offset: int = 0, col_offset: int = 0) -> str:
    """Returns 'product', 'reference', or 'off_wafer' for a given (row, col).

    row_offset/col_offset translate FROM the caller's coordinate space INTO
    the plan's own Die Map numbering before the lookup - the wafer plan's
    row/col (1-indexed, top-left origin) is not the same grid as Accretech's
    (wafer-center-relative, can be negative), see
    NanoZPanel._wafer_plan_offset. Pass 0, 0 (the default) when row/col are
    already in the plan's own space."""
    d = plan.dies.get((row - row_offset, col - col_offset))
    return d["status"] if d else "off_wafer"


def touchdown_slot_exclusions(die_col: int, start_row: int, end_row: int, plan: "WaferPlan",
                              row_offset: int = 0, col_offset: int = 0) -> dict:
    """Slot (1..probe_height, top to bottom of this touchdown) -> exclusion reason,
    or None if that slot lands on a normal product die that should be run.
    die_col/start_row/end_row are in the caller's space; see classify_die."""
    result = {}
    for slot in range(1, plan.probe_height + 1):
        physical_row = start_row + slot - 1
        if physical_row > end_row:
            result[slot] = "past touchdown end"
            continue
        status = classify_die(plan, physical_row, die_col, row_offset, col_offset)
        result[slot] = {"off_wafer": "off wafer", "reference": "reference die",
                        "product": None}[status]
    return result


def wafer_plan_die_grid(plan: "WaferPlan") -> list[dict]:
    """Every on-wafer die (product or reference) as {row, col, status, serial}."""
    return [{"row": r, "col": c, "status": d["status"], "serial": d["serial"]}
           for (r, c), d in sorted(plan.dies.items())]


def wafer_plan_stats(plan: "WaferPlan") -> dict:
    counts = {"product": 0, "reference": 0, "off_wafer": 0}
    for start_row, die_col in plan.touchdowns:
        end_row = start_row + plan.probe_height - 1
        for reason in touchdown_slot_exclusions(die_col, start_row, end_row, plan).values():
            if reason is None:
                counts["product"] += 1
            elif reason == "reference die":
                counts["reference"] += 1
            else:
                counts["off_wafer"] += 1
    return counts


def _build_shot(plan: "WaferPlan", die_col: int, start: int, end: int, ports: list,
                slots_by_port: dict, label: str,
                row_offset: int = 0, col_offset: int = 0) -> dict:
    """Each NanoZ board has two independent chips (0 and 1), each wired to its
    own physical probe-head slot — `slots_by_port[port]` is a {"0": slot_or_None,
    "1": slot_or_None} dict (see BoardIdentity.chip_slots()). A `run <nn>`
    always actuates both chips together (confirmed in the vendor manual, no
    per-chip run command exists), so a board is only excluded from a shot if
    BOTH of its chips land off a normal product die for this touchdown — if
    at least one chip has a real die there, the board still needs to run.
    "chip_reasons" (port -> {"0": reason_or_None, "1": reason_or_None}) records
    the per-chip detail for display/filtering; "board_reasons" (port -> reason
    string, or None if it runs) is the board-level summary, independent of
    whatever the manual excluded_boards toggle grid does to it afterwards.

    die_col/start/end are in the CALLER's coordinate space (e.g. Accretech's)
    and are stored as-is in the returned shot - only the classify_die lookups
    are translated into the plan's own space via row_offset/col_offset, so
    the shot's die_column/td_start_row/td_end_row stay usable for driving
    the physical prober."""
    exclusions = touchdown_slot_exclusions(die_col, start, end, plan, row_offset, col_offset)
    excluded_boards = set()
    board_reasons = {}
    chip_reasons = {}
    for port in ports:
        chip_slots = slots_by_port.get(port) or {}
        per_chip = {}
        for chip in ("0", "1"):
            slot = chip_slots.get(chip)
            if slot is None:
                per_chip[chip] = "no slot assigned"
            else:
                per_chip[chip] = exclusions.get(slot, "slot beyond probe head height")
        chip_reasons[port] = per_chip
        if all(r is not None for r in per_chip.values()):
            excluded_boards.add(port)
            board_reasons[port] = "; ".join(
                f"chip{c}: {r}" for c, r in per_chip.items())
        else:
            board_reasons[port] = None
    return {
        "label": label,
        "excluded_boards": excluded_boards,
        "board_reasons": board_reasons,
        "chip_reasons": chip_reasons,
        "die_column": die_col, "td_start_row": start, "td_end_row": end,
    }


def build_shots_from_windows(plan: "WaferPlan", windows: list, ports: list,
                             slots_by_port: dict,
                             row_offset: int = 0, col_offset: int = 0) -> list[dict]:
    """One shot per (row, col) window - each is a manually-positioned 1-wide x
    probe_height-tall touchdown footprint (e.g. dies the user highlighted on
    the Run tab's wafer map, each imagined as a touchdown's top die), rather
    than the wafer plan's own pre-computed touchdown list. windows are in the
    caller's coordinate space; see _build_shot."""
    shots = []
    for start_row, die_col in windows:
        end = start_row + plan.probe_height - 1
        label = f"Col {die_col} · rows {start_row}-{end} (from selection)"
        shots.append(_build_shot(plan, die_col, start_row, end, ports, slots_by_port, label,
                                 row_offset, col_offset))
    return shots


def active_ports_for_window(plan: "WaferPlan", die_col: int, start_row: int,
                            ports: list, slots_by_port: dict,
                            row_offset: int = 0, col_offset: int = 0) -> list:
    """Ports whose chip(s) land on a real product die within the touchdown
    window starting at start_row in die_col (caller's coordinate space) - the
    same boards Compute Recipe/build_shots_from_windows would leave
    un-excluded for a touchdown anchored here, used by Run Cycle (Active)/
    Pause (Active) to scope to the current position window instead of
    "every connected board"."""
    end_row = start_row + plan.probe_height - 1
    shot = _build_shot(plan, die_col, start_row, end_row, ports, slots_by_port, "",
                       row_offset, col_offset)
    return [p for p in ports if p not in shot["excluded_boards"]]


class NanoZBoard:

    def __init__(self, identity: BoardIdentity, out_queue,
                die_provider: Optional[Callable[[Optional[str]], tuple]] = None,
                env_interval_s: float = 1.0):
        self.identity = identity
        self.port = identity.port
        self.out_queue = out_queue
        # die_provider(chip) -> (row, col); chip is "0"/"1" for a per-chip SPL
        # reading, or None for a board-wide ENV reading.
        self._die_provider = die_provider or (lambda chip: (None, None))
        self.env_interval_s = env_interval_s
        self.ser: Optional[serial.Serial] = None
        self._buffer = bytearray()
        self._thread: Optional[threading.Thread] = None
        self._running = False
        self.spl_count = 0
        self.env_count = 0
        self.last_error = ""

    @property
    def is_running(self) -> bool:
        return self._running

    @property
    def state(self) -> str:
        if self._running and self.last_error:
            return "error"
        if self._running:
            return "connected"
        return "not_connected"

    def connect(self):
        if self.ser is None:
            self.ser = open_serial(self.port)

    def start(self):
        self.last_error = ""
        self.connect()
        try:
            send_ascii(self.ser, "pause")
            time.sleep(0.2)
            self.ser.reset_input_buffer()
        except Exception:
            pass
        self._running = True
        self._thread = threading.Thread(target=self._reader_loop, daemon=True)
        self._thread.start()

    def reconnect(self):
        self.stop()
        self.start()

    def run_cycle(self, cycle: int):
        if self.ser:
            send_ascii(self.ser, f"run {cycle}")

    def pause(self):
        if self.ser:
            send_ascii(self.ser, "pause")

    def request_eeprom(self, addr: int, length: int):
        """Read-only: request a raw non-volatile-memory block (rdeep). The
        response arrives asynchronously as a "kind": "eep" item on
        out_queue, decoded by _handle_eep. Does not actuate anything on the
        board - safe to call at any time, including mid-cycle."""
        if self.ser:
            send_ascii(self.ser, f"rdeep {addr} {length}")

    def send_raw(self, cmd: str):
        if self.ser:
            send_ascii(self.ser, cmd)

    def write_eeprom(self, addr: int, data: bytes):
        """Write a raw block to EEPROM (wreep). DANGEROUS - see the NanoZ_EK
        tab and project notes: only ever call this with a byte-for-byte
        patch of a record that was just read (encode_sequence_patch), never
        a from-scratch record, since large parts of the layout aren't
        understood well enough to safely regenerate. Per the protocol doc,
        the board rejects the write outright (no partial/garbled write) if
        the checksum doesn't match what it computes, or if the full <len>
        bytes don't arrive within ~1s of the command line - the caller
        should still verify afterward with a fresh rdeep, since a rejection
        is only reported as a text error line, not a return value here."""
        if not self.ser:
            return
        cs = 0
        for b in data:
            cs ^= b
        send_ascii(self.ser, f"wreep {len(data)} {cs:04X} {addr}")
        self.ser.write(bytes(data))
        self.ser.flush()

    def stop(self):
        self._running = False
        if self._thread:
            self._thread.join(timeout=2.0)
            self._thread = None
        if self.ser:
            try:
                self.ser.close()
            except Exception:
                pass
        self.ser = None


    def _reader_loop(self):
        next_env = time.time() + self.env_interval_s if self.env_interval_s > 0 else float("inf")
        while self._running:
            try:
                if self.env_interval_s > 0 and time.time() >= next_env:
                    send_ascii(self.ser, "#env?")
                    next_env = time.time() + self.env_interval_s
                line = read_line_bytes(self.ser, self._buffer, timeout_s=0.2)
            except Exception as e:
                self.last_error = str(e)
                time.sleep(0.2)
                continue
            if line is None:
                continue
            self.last_error = ""
            if not line.startswith(b"#") or line.startswith(b"##"):
                self._emit_text(line)
                continue
            sm = SPL_HEADER_RE.match(line)
            if sm:
                self._handle_spl(sm)
                continue
            em = ENV_HEADER_RE.match(line)
            if em:
                self._handle_env(em)
                continue
            eepm = EEP_HEADER_RE.match(line)
            if eepm:
                self._handle_eep(eepm)
                continue
            if SEQ_HEADER_RE.match(line):
                self._emit_text(line)
                continue
            self.out_queue.put({
                "kind": "unrecognized", "board_sn": self.identity.serial_number,
                "port": self.port, "raw": line, "host_timestamp": now_stamp(),
            })

    def _emit_text(self, line: bytes):
        self.out_queue.put({
            "kind": "text", "board_sn": self.identity.serial_number,
            "port": self.port, "text": line.decode(errors="replace"),
            "host_timestamp": now_stamp(),
        })

    def _handle_spl(self, m):
        length_s, cs_s, chip_s, time_s, bfr_s = m.groups()
        length, header_chip, header_time, header_bfr = (
            int(length_s), int(chip_s), int(time_s), int(bfr_s))
        expected_cs = int(cs_s, 16)
        try:
            data = read_exact_from_buffer(self.ser, self._buffer, length, timeout_s=2.0)
        except NanoZError as e:
            self.last_error = str(e)
            return
        try:
            parsed = parse_spl_data(data)
        except Exception as e:
            parsed = {"parse_error": str(e)}
        row, col = self._die_provider(str(header_chip))
        self.spl_count += 1
        self.out_queue.put({
            "kind": "spl", "host_timestamp": now_stamp(),
            "board_sn": self.identity.serial_number, "port": self.port,
            "die_row": row, "die_col": col,
            "header_chip": header_chip, "header_time_ms": header_time,
            "header_bfr": header_bfr, "len": length,
            "checksum_expected": expected_cs,
            **parsed,
        })

    def _handle_env(self, m):
        x_s, length_s, cs_s, time_s, bfr_s = m.groups()
        env_x, length, header_time, header_bfr = (
            int(x_s), int(length_s), int(time_s), int(bfr_s))
        expected_cs = int(cs_s, 16)
        try:
            data = read_exact_from_buffer(self.ser, self._buffer, length, timeout_s=2.0)
        except NanoZError as e:
            self.last_error = str(e)
            return
        try:
            parsed = parse_env_data(data)
        except Exception as e:
            parsed = {"parse_error": str(e)}
        row, col = self._die_provider(None)
        self.env_count += 1
        self.out_queue.put({
            "kind": "env", "host_timestamp": now_stamp(),
            "board_sn": self.identity.serial_number, "port": self.port,
            "die_row": row, "die_col": col,
            "env_x": env_x, "header_time_ms": header_time, "header_bfr": header_bfr,
            "len": length, "checksum_expected": expected_cs,
            **parsed,
        })

    def _handle_eep(self, m):
        length_s, cs_s, addr_s = m.groups()
        length, addr = int(length_s), int(addr_s)
        expected_cs = int(cs_s, 16)
        try:
            data = read_exact_from_buffer(self.ser, self._buffer, length, timeout_s=2.0)
        except NanoZError as e:
            self.last_error = str(e)
            return
        actual_cs = 0
        for b in data:
            actual_cs ^= b
        self.out_queue.put({
            "kind": "eep", "host_timestamp": now_stamp(),
            "board_sn": self.identity.serial_number, "port": self.port,
            "addr": addr, "len": length,
            "checksum_expected": expected_cs, "checksum_actual": actual_cs,
            "checksum_ok": actual_cs == expected_cs,
            "data_hex": data.hex(),
        })
